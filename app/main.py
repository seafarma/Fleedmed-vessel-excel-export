import io
import os
import re
import smtplib
import time
import secrets
from email.message import EmailMessage
from datetime import datetime, date, timedelta
from decimal import Decimal
from typing import List, Dict, Any, Optional, Tuple

import psycopg2
from psycopg2.extras import RealDictCursor
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse, HTMLResponse, RedirectResponse

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL env var is required")

TEMPLATE_PATH = os.getenv("TEMPLATE_PATH", "app/templates/inventory_template 2.03.xlsx")

INVENTORY_SHEET = os.getenv("INVENTORY_SHEET", "Inventory")
VESSEL_INFO_SHEET = os.getenv("VESSEL_INFO_SHEET", "Vessel Information")
UPLOAD_SHEET = os.getenv("UPLOAD_SHEET", "Upload")
STORAGE_SHEET = os.getenv("STORAGE_SHEET", "Storage")

MAX_ROWS = int(os.getenv("MAX_ROWS", "3000"))

SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_USE_TLS = os.getenv("SMTP_USE_TLS", "true").lower() == "true"
FROM_EMAIL = os.getenv("FROM_EMAIL", SMTP_USER or "no-reply@example.com")
FROM_NAME = os.getenv("FROM_NAME", "Inventory Export")

app = FastAPI()

WAIT_HTML = """
<!doctype html>
<html>
  <head>
    <meta charset="utf-8" />
    <title>Preparing download…</title>
    <meta http-equiv="refresh" content="1">
    <style>
      body { font-family: Arial, sans-serif; padding: 24px; }
    </style>
  </head>
  <body>
    Preparing your Excel file… this page will refresh automatically.
  </body>
</html>
"""


# Vessel Information template positions (based on your latest exported file)
CELL_NEXT_RESUPPLY = "I4"        # label at I3
CELL_RESUPPLY_AGREEMENT = "I6"   # label at I5
CELL_RESUPPLY_RATE = "I8"        # label at I7
CELL_EXPIRATION_MONTHS = "I11"   # label around I9
CELL_NOTES = "N4"                # label at N3

# Certificate list location in your template (label "Certificate expiry date" at K3)
CERT_PACK_COL = "K"
CERT_STATUS_COL = "L"
CERT_START_ROW = 4
CERT_MAX_ROWS = 27  # fills K4:L30


def db_conn():
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def fetch_one(sql: str, params: tuple) -> Optional[Dict[str, Any]]:
    with db_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return cur.fetchone()


def fetch_all(sql: str, params: tuple) -> List[Dict[str, Any]]:
    with db_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return cur.fetchall()


def get_excel_row(excel_id: str) -> Dict[str, Any]:
    row = fetch_one(
        "SELECT excel_id, vessel_id, export_token FROM vessel_excel WHERE excel_id=%s",
        (excel_id,),
    )
    if not row:
        raise HTTPException(status_code=404, detail="excel_id not found")
    return row


def get_excel_row_retry(
    excel_id: str, max_wait_s: float = 6.0, step_s: float = 0.25
) -> Optional[Dict[str, Any]]:
    deadline = time.time() + max_wait_s
    while True:
        row = fetch_one(
            "SELECT excel_id, vessel_id, export_token FROM vessel_excel WHERE excel_id=%s",
            (excel_id,),
        )
        if row:
            return row
        if time.time() >= deadline:
            return None
        time.sleep(step_s)


def validate_token(excel_row: Dict[str, Any], token: str):
    expected = (excel_row.get("export_token") or "").strip()
    tok = (token or "").strip()
    if not expected or not secrets.compare_digest(tok, expected):
        raise HTTPException(status_code=403, detail="invalid token")


def to_bool(v) -> Optional[bool]:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    s = str(v).strip().lower()
    if s in ("true", "t", "yes", "y", "1", "on"):
        return True
    if s in ("false", "f", "no", "n", "0", "off"):
        return False
    return None


def yesno_or_blank(v) -> str:
    b = to_bool(v)
    if b is None:
        return ""
    return "Yes" if b else "No"


def parse_date_any(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    try:
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except Exception:
        pass
    try:
        if len(s) >= 10 and s[2] == "-" and s[5] == "-":
            return datetime.strptime(s[:10], "%d-%m-%Y").date()
    except Exception:
        pass
    return None


def txt(v) -> str:
    return "" if v is None else str(v)


def num(v) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, Decimal):
        return float(v)
    try:
        return float(v)
    except Exception:
        return None


def find_sheet_fuzzy(wb, wanted: str):
    w = wanted.strip().lower()
    for n in wb.sheetnames:
        if n.strip().lower() == w:
            return wb[n]
    for n in wb.sheetnames:
        if w in n.strip().lower():
            return wb[n]
    return None


def try_get_vessel_json(vessel_id: str) -> Dict[str, Any]:
    # Prefer enriched view; fallback vessels
    try:
        row = fetch_one(
            "SELECT to_jsonb(t) AS j FROM vw_vessels_enriched t WHERE t.vessel_id=%s",
            (vessel_id,),
        )
        if row and row.get("j"):
            return row["j"]
    except Exception:
        pass

    row2 = fetch_one(
        "SELECT to_jsonb(v) AS j FROM vessels v WHERE v.vessel_id=%s",
        (vessel_id,),
    )
    if row2 and row2.get("j"):
        return row2["j"]
    return {}


def looks_like_hex_id(s: str) -> bool:
    s = (s or "").strip()
    # your IDs look like "d11259bd" (8 hex). Allow 8+ hex.
    return bool(re.fullmatch(r"[0-9a-fA-F]{8,}", s))


def lookup_vessel_category_name(cat_id: str) -> str:
    if not cat_id:
        return ""
    candidates = [
        ("vessel_categories", "category_name", "vessel_category_id"),
        ("vessel_categories", "category_name", "category_id"),
        ("vessel_category", "category_name", "vessel_category_id"),
        ("vessel_category", "category_name", "category_id"),
    ]
    for table, namecol, idcol in candidates:
        try:
            row = fetch_one(
                f"SELECT {namecol} AS n FROM {table} WHERE {idcol}=%s LIMIT 1",
                (cat_id,),
            )
            if row and row.get("n"):
                return str(row["n"])
        except Exception:
            continue
    return ""


def resolve_vessel_category(vessel: Dict[str, Any]) -> str:
    """
    Fix for your "Vessel Category is empty":
    sometimes vessel already contains the NAME (not the ID).
    If it looks like an ID -> lookup; otherwise use it directly.
    """
    raw = (
        vessel.get("vessel_category_name")
        or vessel.get("vessel_category")
        or vessel.get("vessel_category_id")
        or ""
    )
    s = str(raw or "").strip()
    if not s:
        return ""

    if looks_like_hex_id(s):
        return lookup_vessel_category_name(s) or ""
    # already a name
    return s


def lookup_flag_name(flag_id_or_name: str) -> str:
    if not flag_id_or_name:
        return ""
    s = str(flag_id_or_name).strip()
    # If it's not an ID-looking value, likely already a display name (e.g. "Sweden")
    if s and not looks_like_hex_id(s):
        return s

    candidates = [
        ("entities", "entity_name", "entity_id"),
        ("entities", "entity_name", "entity_code"),
        ("entity", "entity_name", "entity_id"),
        ("vessel_flags", "entity_name", "vessel_flag_id"),
        ("flags", "entity_name", "flag_id"),
    ]
    for table, namecol, idcol in candidates:
        try:
            row = fetch_one(
                f"SELECT {namecol} AS n FROM {table} WHERE {idcol}=%s LIMIT 1",
                (s,),
            )
            if row and row.get("n"):
                return str(row["n"])
        except Exception:
            continue
    return s


def build_filename(vessel: Dict[str, Any], vessel_id: str) -> str:
    raw_name = vessel.get("vessel_name") or ""
    raw_imo = vessel.get("vessel_IMO") or ""

    if vessel_id and (not raw_name or not raw_imo):
        try:
            row = fetch_one(
                'SELECT vessel_name, "vessel_IMO" AS imo FROM vessels WHERE vessel_id=%s',
                (vessel_id,),
            ) or {}
            raw_name = raw_name or (row.get("vessel_name") or "")
            raw_imo = raw_imo or (row.get("imo") or "")
        except Exception:
            pass

    vessel_name = str(raw_name or "Vessel").replace("/", "-").strip()
    imo = str(raw_imo or "").strip()
    date_str = datetime.now().strftime("%d-%m-%Y")

    if imo:
        return f"Medical Inventory List - {vessel_name} - {imo} - {date_str}.xlsx"
    return f"Medical Inventory List - {vessel_name} - {date_str}.xlsx"


def get_export_rows(excel_id: str) -> List[Dict[str, Any]]:
    """
    NOTE: we SELECT item_classification_export from the view for Inventory N/M/C/D/F/O,
    but we DO NOT write it to Upload sheet (since you removed that column).
    """
    return fetch_all(
        """
        SELECT
          v.vessel_item_id,
          v.vessel_id,
          v.storage_id,
          v.storage_display,
          v.item_id,
          v.pack_id,
          v.category_id,
          v.item_barcode,
          v.vessel_item_name,
          v.vessel_item_law_code,
          v.vessel_item_quantity,
          v.totalitem_qty_sql,
          v.certificate_qty_sql,
          v.vessel_item_expiration_date,
          v.pack_name,
          v.item_classification_export
        FROM vw_vessel_excel_items v
        WHERE v.excel_id=%s
        ORDER BY
          COALESCE(v.pack_name,'') ASC,
          COALESCE(v.vessel_item_law_code,'') ASC,
          COALESCE(v.storage_display,'') ASC,
          COALESCE(v.vessel_item_name,'') ASC
        """,
        (excel_id,),
    )


def get_vessel_certificates_latest_by_pack(vessel_id: str) -> List[Dict[str, Any]]:
    """
    All packs linked to vessel via vessel_certifications,
    latest certificate_end_date per pack.
    """
    try:
        return fetch_all(
            """
            SELECT
              vc.pack_id,
              p.pack_name,
              MAX(vc.certificate_end_date)::date AS last_certificate_end_date
            FROM vessel_certifications vc
            LEFT JOIN packs p ON p.pack_id = vc.pack_id
            WHERE vc.vessel_id = %s
            GROUP BY vc.pack_id, p.pack_name
            ORDER BY COALESCE(p.pack_name,'') ASC
            """,
            (vessel_id,),
        )
    except Exception:
        return []


def compute_next_resupply(cert_rows: List[Dict[str, Any]]) -> Optional[date]:
    """
    earliest (end_date - 30) across NON-expired certs.
    If none valid -> blank (this matches your earlier rule).
    """
    today = date.today()
    candidates: List[date] = []
    for r in cert_rows or []:
        end_d = parse_date_any(r.get("last_certificate_end_date"))
        if end_d and end_d >= today:
            candidates.append(end_d - timedelta(days=30))
    return min(candidates) if candidates else None


def get_vessel_storages(vessel_id: str) -> List[Tuple[str, str]]:
    """
    Fill Storage sheet: Storage Name + storage_id, sorted by name.
    Uses vessel_storages.
    """
    if not vessel_id:
        return []
    try:
        rows = fetch_all(
            """
            SELECT
              storage_id,
              COALESCE(
                NULLIF(storage_display,''),
                NULLIF(storage_name,''),
                NULLIF(storage_location,''),
                storage_id
              ) AS storage_name
            FROM vessel_storages
            WHERE vessel_id=%s
            """,
            (vessel_id,),
        )
    except Exception:
        return []

    storages: List[Tuple[str, str]] = []
    seen: set = set()
    for r in rows or []:
        sid = (r.get("storage_id") or "").strip()
        name = (r.get("storage_name") or "").strip()
        if not sid:
            continue
        if sid in seen:
            continue
        seen.add(sid)
        storages.append((name or sid, sid))

    storages.sort(key=lambda x: (x[0] or "").lower())
    return storages


def storage_map_from_list(storages: List[Tuple[str, str]]) -> Dict[str, str]:
    return {sid: name for (name, sid) in storages if sid and name}


def excel_quote_sheet(name: str) -> str:
    # Always quote to be safe
    return "'" + str(name).replace("'", "''") + "'"


def apply_storage_dropdown(inv_ws, storage_sheet_name: str, n_storages: int):
    """
    Inventory!Storage column dropdown from Storage sheet column A (Storage Name).
    """
    if n_storages <= 0:
        return
    end_row = 1 + n_storages  # header at row 1
    sheet_ref = excel_quote_sheet(storage_sheet_name)
    formula = f"={sheet_ref}!$A$2:$A${end_row}"

    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.errorTitle = "Invalid storage"
    dv.error = "Select a storage from the list."
    dv.showErrorMessage = True

    inv_ws.add_data_validation(dv)
    dv.add(f"A2:A{MAX_ROWS}")


def fill_storage_sheet(wb, storages: List[Tuple[str, str]]):
    ws = find_sheet_fuzzy(wb, STORAGE_SHEET)
    if ws is None:
        ws = wb.create_sheet(STORAGE_SHEET)
        ws["A1"].value = "Storage Name"
        ws["B1"].value = "storage_id"

    # Clear existing values (keep formatting)
    for r in range(2, 1000):
        ws.cell(r, 1).value = None
        ws.cell(r, 2).value = None

    # Fill
    for i, (name, sid) in enumerate(storages, start=2):
        ws.cell(i, 1).value = name
        ws.cell(i, 2).value = sid

    return ws


def parse_item_classification_to_flags(v: Any) -> Tuple[bool, bool, bool, bool, bool, bool]:
    """
    Returns N, M, C, D, F, O booleans from item_classification_export.
    Supports:
      - codes: "N", "N,M", "NMCDFO"
      - keywords: narc / malar / cool/cold/fridge / danger/dg / fem / oxygen/o2
    """
    s = "" if v is None else str(v).strip()
    if not s:
        return (False, False, False, False, False, False)

    s_up = s.upper()
    compact = re.sub(r"[\s,;/+\-|]+", "", s_up)

    if compact and re.fullmatch(r"[NMCDFO]{1,10}", compact):
        flags = set(compact)
        return (
            "N" in flags,
            "M" in flags,
            "C" in flags,
            "D" in flags,
            "F" in flags,
            "O" in flags,
        )

    sl = s.lower()
    n = "narc" in sl
    m = "malar" in sl
    c = ("cool" in sl) or ("cold" in sl) or ("fridge" in sl)
    d = ("danger" in sl) or ("dg" in sl) or ("hazmat" in sl)
    f = "fem" in sl
    o = ("oxygen" in sl) or ("o2" in sl) or ("oxy" in sl) or (sl.strip() == "o")
    return (n, m, c, d, f, o)


def fill_vessel_information_sheet(
    wb, vessel: Dict[str, Any], excel_row: Dict[str, Any], cert_rows: List[Dict[str, Any]]
):
    ws = find_sheet_fuzzy(wb, VESSEL_INFO_SHEET)
    if ws is None:
        ws = wb.create_sheet(VESSEL_INFO_SHEET)

    # Base vessel row (if available) may contain additional fields
    vessel_id = (excel_row or {}).get("vessel_id") or ""
    if vessel_id:
        row = fetch_one(
            "SELECT to_jsonb(v) AS j FROM vessels v WHERE v.vessel_id=%s",
            (vessel_id,),
        )
        if row and row.get("j"):
            base_vessel = row["j"]
            vessel = {**base_vessel, **(vessel or {})}

    # Left block (same as your current template)
    ws["A4"].value = txt(vessel.get("company_name") or "")
    ws["C4"].value = txt(vessel.get("vessel_name") or "")
    ws["E4"].value = yesno_or_blank(vessel.get("malaria_area"))
    ws["G4"].value = yesno_or_blank(vessel.get("mfag"))

    ws["A6"].value = txt(vessel.get("vessel_contact_name") or "")
    ws["C6"].value = txt(vessel.get("vessel_IMO") or "")
    ws["E6"].value = yesno_or_blank(vessel.get("female_onboard"))

    ws["A8"].value = txt(vessel.get("vessel_contact_email") or vessel.get("email") or "")
    ws["C8"].value = txt(lookup_flag_name(vessel.get("vessel_flag_name") or vessel.get("vessel_flag") or ""))
    ws["E8"].value = yesno_or_blank(vessel.get("dangerous_good"))

    ws["A10"].value = txt(vessel.get("vessel_contact_phone") or "")
    ws["C10"].value = txt(vessel.get("vessel_crew_size") or "")
    ws["E10"].value = yesno_or_blank(vessel.get("narcotics"))

    purchasing = (
        vessel.get("purchasing_email")
        or vessel.get("purchasing_mail")
        or vessel.get("purchasing")
        or ""
    )
    if not purchasing:
        purchasing = vessel.get("vessel_contact_email") or vessel.get("email") or ""
    ws["A12"].value = txt(purchasing)

    ws["C12"].value = txt(resolve_vessel_category(vessel))
    ws["E12"].value = yesno_or_blank(vessel.get("medical_oxygen"))

    # Right block (your template positions)
    next_resupply = compute_next_resupply(cert_rows)
    if next_resupply:
        ws[CELL_NEXT_RESUPPLY].value = next_resupply
        ws[CELL_NEXT_RESUPPLY].number_format = "dd-mm-yyyy"
    else:
        ws[CELL_NEXT_RESUPPLY].value = ""

    agreement = vessel.get("vessel_subscription_type") or ""
    ws[CELL_RESUPPLY_AGREEMENT].value = txt(agreement)

    rr = vessel.get("resupply_rate")
    ws[CELL_RESUPPLY_RATE].value = rr if rr not in (None, "") else ""

    em = vessel.get("expiration_months")
    ws[CELL_EXPIRATION_MONTHS].value = em if em not in (None, "") else ""

    # Notes (your template uses N4)
    ws[CELL_NOTES].value = txt(vessel.get("vessel_notes") or "")

    # Certificates list (K4/L4 downward)
    # Clear old values first
    for r in range(CERT_START_ROW, CERT_START_ROW + CERT_MAX_ROWS):
        ws[f"{CERT_PACK_COL}{r}"].value = None
        ws[f"{CERT_STATUS_COL}{r}"].value = None
        ws[f"{CERT_STATUS_COL}{r}"].number_format = "General"

    today = date.today()
    for idx, cr in enumerate(cert_rows[:CERT_MAX_ROWS]):
        r = CERT_START_ROW + idx
        pack_name = (cr.get("pack_name") or "").strip() or (cr.get("pack_id") or "")
        end_d = parse_date_any(cr.get("last_certificate_end_date"))

        ws[f"{CERT_PACK_COL}{r}"].value = txt(pack_name)

        if end_d and end_d >= today:
            ws[f"{CERT_STATUS_COL}{r}"].value = end_d
            ws[f"{CERT_STATUS_COL}{r}"].number_format = "dd-mm-yyyy"
        elif end_d:
            ws[f"{CERT_STATUS_COL}{r}"].value = "Expired"
        else:
            ws[f"{CERT_STATUS_COL}{r}"].value = ""


def fill_inventory_sheet(inv_ws, rows: List[Dict[str, Any]]):
    # map headers (row 1)
    headers: Dict[str, int] = {}
    for c in range(1, inv_ws.max_column + 1):
        v = inv_ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip().lower()] = c

    def col(name: str) -> int:
        c = headers.get(name.lower())
        if not c:
            raise HTTPException(status_code=500, detail=f"Inventory template missing column: {name}")
        return c

    c_storage = col("Storage")
    c_article = col("Article No.")
    c_item = col("Item name")
    c_qty = col("Quantity")
    c_total = col("Total quantity")
    c_cert = col("Certificate qty")
    c_exp = col("Expiry date")
    c_law = col("Law code")
    c_pack = col("Pack name")

    # Required classification columns in your template
    c_n = col("N")
    c_m = col("M")
    c_c = col("C")
    c_d = col("D")
    c_f = col("F")
    c_o = col("O")

    # Clear values (keep formatting)
    for r in range(2, MAX_ROWS + 1):
        for cidx in (c_storage, c_article, c_item, c_qty, c_total, c_cert, c_exp, c_law, c_pack, c_n, c_m, c_c, c_d, c_f, c_o):
            inv_ws.cell(r, cidx).value = None

    # Write rows
    out_r = 2
    for rr in rows:
        if out_r > MAX_ROWS:
            break

        storage = (rr.get("storage_display") or "").strip() or "Inbound storage"
        barcode = (rr.get("item_barcode") or "").strip() or "Extra"

        inv_ws.cell(out_r, c_storage).value = storage
        inv_ws.cell(out_r, c_article).value = barcode
        inv_ws.cell(out_r, c_item).value = rr.get("vessel_item_name") or ""

        q = num(rr.get("vessel_item_quantity")) or 0
        inv_ws.cell(out_r, c_qty).value = q

        inv_ws.cell(out_r, c_total).value = num(rr.get("totalitem_qty_sql")) or 0
        inv_ws.cell(out_r, c_cert).value = num(rr.get("certificate_qty_sql")) or 0

        dte = parse_date_any(rr.get("vessel_item_expiration_date"))
        if dte:
            inv_ws.cell(out_r, c_exp).value = dte
            inv_ws.cell(out_r, c_exp).number_format = "mm-yyyy"
        else:
            inv_ws.cell(out_r, c_exp).value = None

        inv_ws.cell(out_r, c_law).value = rr.get("vessel_item_law_code") or ""
        inv_ws.cell(out_r, c_pack).value = rr.get("pack_name") or ""

        n, m, c, d, f, o = parse_item_classification_to_flags(rr.get("item_classification_export"))
        inv_ws.cell(out_r, c_n).value = "N" if n else None
        inv_ws.cell(out_r, c_m).value = "M" if m else None
        inv_ws.cell(out_r, c_c).value = "C" if c else None
        inv_ws.cell(out_r, c_d).value = "D" if d else None
        inv_ws.cell(out_r, c_f).value = "F" if f else None
        inv_ws.cell(out_r, c_o).value = "O" if o else None

        out_r += 1


def fill_upload_sheet(up_ws, rows: List[Dict[str, Any]]):
    """
    IMPORTANT: You removed item_classification_export from Upload.
    This function does NOT write it.
    """
    headers: Dict[str, int] = {}
    for c in range(1, up_ws.max_column + 1):
        v = up_ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip().lower()] = c

    def setv(r: int, h: str, v):
        c = headers.get(h.lower())
        if c:
            up_ws.cell(r, c).value = v

    # Clear rows (keep formatting)
    for r in range(2, MAX_ROWS + 1):
        for c in range(1, up_ws.max_column + 1):
            up_ws.cell(r, c).value = None

    out_r = 2
    for rr in rows:
        if out_r > MAX_ROWS:
            break

        storage = (rr.get("storage_display") or "").strip() or "Inbound storage"
        barcode = (rr.get("item_barcode") or "").strip() or "123"
        dte = parse_date_any(rr.get("vessel_item_expiration_date"))

        setv(out_r, "vessel_item_id", rr.get("vessel_item_id") or "")
        setv(out_r, "vessel_id", rr.get("vessel_id") or "")
        setv(out_r, "storage_id", rr.get("storage_id") or "")
        setv(out_r, "item_id", rr.get("item_id") or "")
        setv(out_r, "pack_id", rr.get("pack_id") or "")
        setv(out_r, "category_id", rr.get("category_id") or "")
        setv(out_r, "storage_display", storage)
        setv(out_r, "item_barcode", barcode)
        setv(out_r, "vessel_item_name", rr.get("vessel_item_name") or "")
        setv(out_r, "vessel_item_quantity", num(rr.get("vessel_item_quantity")) or 0)
        setv(out_r, "totalitem_qty_sql", num(rr.get("totalitem_qty_sql")) or 0)
        setv(out_r, "certificate_qty_sql", num(rr.get("certificate_qty_sql")) or 0)
        setv(out_r, "vessel_item_expiration_date", dte or "")
        setv(out_r, "vessel_item_law_code", rr.get("vessel_item_law_code") or "")
        setv(out_r, "pack_name", rr.get("pack_name") or "")

        # set number format if date column exists
        if dte:
            c = headers.get("vessel_item_expiration_date")
            if c:
                up_ws.cell(out_r, c).number_format = "dd-mm-yyyy"

        out_r += 1


def build_workbook_bytes(
    excel_row: Dict[str, Any],
    rows: List[Dict[str, Any]],
    vessel: Dict[str, Any],
    cert_rows: List[Dict[str, Any]],
    storages: List[Tuple[str, str]],
) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)

    # Fill Storage sheet
    ws_storage = fill_storage_sheet(wb, storages)

    # Inventory
    ws_inv = find_sheet_fuzzy(wb, INVENTORY_SHEET)
    if ws_inv is None:
        raise HTTPException(status_code=500, detail=f"Sheet '{INVENTORY_SHEET}' not found in template")
    fill_inventory_sheet(ws_inv, rows)

    # Apply dropdown validation on Inventory Storage column
    apply_storage_dropdown(ws_inv, ws_storage.title, len(storages))

    # Upload
    ws_up = find_sheet_fuzzy(wb, UPLOAD_SHEET)
    if ws_up is not None:
        fill_upload_sheet(ws_up, rows)

    # Vessel Information
    fill_vessel_information_sheet(wb, vessel, excel_row, cert_rows)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@app.get("/health")
def health():
    return {"ok": True}


@app.get("/download")
def download(excel_id: str, token: str):
    excel_row = get_excel_row_retry(excel_id, max_wait_s=6.0, step_s=0.25)
    if not excel_row:
        return HTMLResponse(WAIT_HTML, status_code=200)

    validate_token(excel_row, token)
    return RedirectResponse(
        url=f"/download_file?excel_id={excel_id}&token={token}",
        status_code=302,
    )


@app.get("/download_file")
def download_file(excel_id: str, token: str):
    excel_row = get_excel_row_retry(excel_id, max_wait_s=6.0, step_s=0.25)
    if not excel_row:
        raise HTTPException(status_code=404, detail="excel_id not found")

    validate_token(excel_row, token)

    vessel_id = excel_row["vessel_id"]
    vessel = try_get_vessel_json(vessel_id)

    # Certificates per pack (latest only)
    cert_rows = get_vessel_certificates_latest_by_pack(vessel_id)

    # Storages list for Storage sheet + dropdown
    storages = get_vessel_storages(vessel_id)
    storage_map = storage_map_from_list(storages)

    rows = get_export_rows(excel_id)

    # Fix empty storage_display via vessel_storages mapping
    if storage_map:
        for rr in rows:
            if not (rr.get("storage_display") or "").strip():
                sid = (rr.get("storage_id") or "").strip()
                if sid and storage_map.get(sid):
                    rr["storage_display"] = storage_map[sid]

    # Skip qty 0 items (Inventory + Upload)
    filtered_rows: List[Dict[str, Any]] = []
    for rr in rows:
        q = num(rr.get("vessel_item_quantity"))
        q = 0 if q is None else q
        if q == 0:
            continue
        filtered_rows.append(rr)

    content = build_workbook_bytes(excel_row, filtered_rows, vessel, cert_rows, storages)
    filename = build_filename(vessel, vessel_id)

    safe_name = filename.replace('"', "").replace("\n", " ").replace("\r", " ")
    headers = {
        "Content-Disposition": f'attachment; filename="{safe_name}"',
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/email")
def email(excel_id: str, token: str, to_email: str):
    excel_row = get_excel_row_retry(excel_id, max_wait_s=6.0, step_s=0.25)
    if not excel_row:
        raise HTTPException(status_code=404, detail="excel_id not found")

    validate_token(excel_row, token)

    vessel_id = excel_row["vessel_id"]
    vessel = try_get_vessel_json(vessel_id)
    cert_rows = get_vessel_certificates_latest_by_pack(vessel_id)

    storages = get_vessel_storages(vessel_id)
    storage_map = storage_map_from_list(storages)

    rows = get_export_rows(excel_id)

    if storage_map:
        for rr in rows:
            if not (rr.get("storage_display") or "").strip():
                sid = (rr.get("storage_id") or "").strip()
                if sid and storage_map.get(sid):
                    rr["storage_display"] = storage_map[sid]

    filtered_rows: List[Dict[str, Any]] = []
    for rr in rows:
        q = num(rr.get("vessel_item_quantity"))
        q = 0 if q is None else q
        if q == 0:
            continue
        filtered_rows.append(rr)

    content = build_workbook_bytes(excel_row, filtered_rows, vessel, cert_rows, storages)
    filename = build_filename(vessel, vessel_id)

    subject = f"Medical Inventory List - {vessel.get('vessel_name','')}".strip()
    body = "Please find the medical inventory list attached."

    msg = EmailMessage()
    msg["From"] = f"{FROM_NAME} <{FROM_EMAIL}>"
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)
    msg.add_attachment(
        content,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )

    if SMTP_USE_TLS:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls()
            if SMTP_USER:
                s.login(SMTP_USER, SMTP_PASSWORD)
            s.send_message(msg)
    else:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            if SMTP_USER:
                s.login(SMTP_USER, SMTP_PASSWORD)
            s.send_message(msg)

    return {"ok": True, "sent_to": to_email}


@app.get("/")
def index():
    return HTMLResponse("<h3>Vessel Excel Export</h3><p>Use /download?excel_id=...&token=...</p>")
